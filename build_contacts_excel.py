import csv, re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter, defaultdict

# ── Start dates ────────────────────────────────────────────────────────────────
START_DATES = {
    'Nan Houn':'01-Jun-2016','Nuthida Han':'05-Dec-2017','Ratana Khy':'01-Oct-2018',
    'Vanthen Kim':'01-Nov-2018','Nhoem David':'01-Jun-2019','Davit Mao':'01-Jun-2019',
    'Sila Kun':'01-Jun-2019','Vattey Chhun':'01-Sep-2019','Hin Loeb':'01-Mar-2020',
    'Sopheap Puth':'21-Jul-2020','Chamroeun Muon':'01-Aug-2014','Rina Nov':'23-Feb-2015',
    'Sophea Mak':'01-Aug-2014','Sokuntheary Sim':'01-Feb-2016','Sreileak Rous':'01-Feb-2016',
    'Lieng Huon':'11-May-2016','Barang Soeurm':'01-Jun-2016','Bopha Mom':'01-Jun-2016',
    'Sombo Ros':'01-Jun-2016','Kim Sopheap':'01-Jun-2016','Karano Chhuon':'20-Jul-2016',
    'Matthew Seng':'01-Aug-2016','Sophy Chek':'01-Aug-2016','Phin Teb':'01-Feb-2017',
    'Seyla Nga':'11-Apr-2017','Champa Toch':'06-Jun-2017','Kolap Touch':'06-Jun-2017',
    'Sreymom Lim':'03-Jan-2018','Sopheap Heang':'01-Mar-2018','Sreypeou Hout':'19-Mar-2018',
    'Srey Roth Outh':'24-Jul-2018','Tola Khem':'15-Oct-2018','Hou Sokcheat':'22-Jan-2019',
    'Pech Sang':'01-Apr-2019','Parigna Souem':'01-May-2019','Hean Luos':'01-Jun-2019',
    'Sopheap Ngat':'12-Aug-2019','Ya Hoeun':'01-Jan-2020','Thon Tep':'01-Jan-2020',
    'Rin Nga':'01-Mar-2020','Makara Tes':'01-Apr-2020','Som Chhang':'01-May-2020',
    'Srytom Pogn':'16-Jun-2020','Ousaphea Lim':'01-Sep-2020','Sreymom Thorng':'01-Nov-2020',
    'Kimbuoy Pheng':'12-Jan-2021','Panha Neang':'01-Feb-2014','Rany Mom':'15-Jun-2015',
    'Loy Sambo Hout':'01-Feb-2015','Ryna Mom':'01-Feb-2022','Matthias Lendi':'15-Aug-2015',
    'Martin Strupler':'01-Apr-2021','Thavy Tham':'01-Apr-2021','Hing Hoeun':'01-Feb-2021',
    'Chrach Chen':'09-Feb-2021','Seangly Leng':'20-Apr-2021','Longsamnieng Pol':'01-Apr-2021',
    'Makara Seng':'01-Mar-2021','Sreyleak So':'01-Feb-2021','Khemry Men':'01-Jun-2021',
    'Sopheak Vat':'08-Jun-2021','Mom Kalic':'22-Jun-2021','Rothana Ros':'01-Nov-2021',
    'Kongkea Kouk':'01-Aug-2022','Saroeun Kim':'01-Jun-2022','Sreymom Touch':'15-Sep-2021',
    'Aline Barbosa':'04-Jan-2021','Robson Barbosa':'04-Jan-2021','Steffi Lendi':'15-Aug-2015',
    'Bethany Roach':'07-Jan-2017','Eddie Roach':'07-Jan-2017','Stephanie Shelow':'28-Feb-2019',
    'ND Strupler':'20-Aug-2013','Simone Strupler':'01-Jan-2019','Sophal Strupler':'10-Jan-2013',
    'Vicheth Song':'15-Sep-2021','Nirorn Sok':'01-Sep-2021','Kanha Un':'01-Dec-2021',
    'Sakorun Pok':'25-Nov-2022','Sothiery Un':'15-Mar-2022','Khunhy Thoeun':'12-Apr-2022',
    'Kim Sorn Soeuth':'05-Apr-2022','Sreypich Chhean':'02-Apr-2022','Soktheavy Tann':'01-Jun-2022',
    'Sovansreyreach Tat':'01-Nov-2023','Sarath Sok':'04-May-2022','Noya Men':'01-Aug-2024',
    'Rongroeung Chamroeun':'18-Oct-2022','Neaksen Sok':'01-Feb-2023','Dana Thy':'01-Aug-2023',
    'Linet Un':'01-Mar-2023','Rachana San':'11-Jul-2023','Vann Rotha':'20-Jul-2023',
    'Sotheany Ngeun':'01-Jan-2024','Sambath Song':'01-Nov-2023','Jane Meas':'16-Dec-2023',
    'Seav Ey Khean':'01-Apr-2024','Phat Sout':'17-Nov-2024','Pichey Ken':'11-Jun-2024',
    'Sengly Muol':'04-Oct-2024','Sotheavy Ngen':'03-Sep-2024','Boromey Hom':'16-Aug-2024',
    'Su Sey':'12-Aug-2024','Sreyleak Orn':'04-Oct-2024','Ranaomi Heak':'06-Dec-2024',
    'Thea Rith Keo':'11-Feb-2025','Vutha Kry':'01-Jan-2025','Sa Soueyvan':'01-Mar-2025',
    'Sodalin Thangham':'01-Mar-2025','Kanal Tensok':'01-Feb-2025','Sideth Duch':'18-Mar-2025',
    'Florin Flammer':'11-Jun-2025','Noa Flammer':'15-Jun-2025','Elea Weber':'19-Sep-2025',
    'Sopheaktra Mao':'01-Oct-2025','Chanthat Leat':'01-Oct-2025','Champa Vong':'01-Oct-2025',
    'David Piseth':'01-Nov-2025','Sen Meta':'11-Nov-2025','Kropumkannika Gnorng':'11-Nov-2025',
    'Sreynouch Khit':'11-Nov-2025','Vichea Khen':'12-Nov-2025','Somun Yem':'24-Feb-2026',
    'Savann Sun':'17-Mar-2026','Savry Klot':'01-Apr-2026','Sotheavy Thea':'01-Apr-2026',
    'Vivian Stumpf':'21-Apr-2026','Rayu Roeun':'28-Apr-2026','Seavva Han':'01-May-2026',
}

# ── Birthdays (from calendar, format DD-Mon) ───────────────────────────────────
BIRTHDAYS = {
    # Aug
    'Sophea Mak':'27-Aug','Somun Yem':'28-Aug',
    # Wait — Aug 2026 calendar: Mak Sophea = Jul 27? Let me re-check.
    # Calendar Aug 2026 week 1 shows Mon Jul 27 = Mak Sophea. So it's Jul 27.
}

# Actually re-doing this properly — calendar shows Aug 2026 view with Jul 27 in week 1
BIRTHDAYS = {
    'Sophea Mak':           '27-Jul',
    'Somun Yem':            '28-Jul',  # Yem Somoun
    'Sophal Strupler':      '01-Aug',
    'Tola Khem':            '03-Aug',
    'Jane Meas':            '15-Aug',
    'Sokuntheary Sim':      '17-Aug',
    'Sombo Ros':            '23-Aug',
    'Thea Rith Keo':        '02-Sep',
    'Thavy Tham':           '03-Sep',
    'Rany Mom':             '07-Sep',
    'Sambath Song':         '07-Sep',
    'Robson Barbosa':       '10-Sep',
    'Eddie Roach':          '16-Sep',
    'Sreyleak So':          '22-Sep',
    'Ranaomi Heak':         '25-Sep',
    'Sotheany Ngeun':       '25-Sep',
    'Hean Luos':            '01-Oct',
    'Ratana Khy':           '05-Oct',
    'Nuthida Han':          '05-Oct',
    'Kim Sopheap':          '05-Oct',
    'Thon Tep':             '05-Oct',
    'Champa Toch':          '05-Oct',
    'Kolap Touch':          '05-Oct',
    'Boromey Hom':          '06-Oct',
    'Seavva Han':           '07-Oct',
    'Srytom Pogn':          '07-Oct',
    'Simone Strupler':      '09-Oct',
    'Panha Neang':          '10-Oct',
    'Kanal Tensok':         '13-Oct',
    'Rothana Ros':          '15-Oct',
    'Khemry Men':           '17-Oct',
    'Makara Tes':           '17-Oct',
    'Vattey Chhun':         '24-Oct',
    'Soktheavy Tann':       '30-Oct',
    'Dana Thy':             '02-Nov',
    'Noa Flammer':          '04-Nov',
    'Chamroeun Muon':       '08-Nov',
    'Su Sey':               '12-Nov',
    'Phin Teb':             '12-Nov',
    'Hin Loeb':             '21-Nov',
    'Seangly Leng':         '25-Nov',
    'Sen Meta':             '02-Dec',
    'Sa Soueyvan':          '03-Dec',
    'Sovansreyreach Tat':   '04-Dec',
    'Sengly Muol':          '07-Dec',
    'Sreileak Rous':        '12-Dec',
    'Rongroeung Chamroeun': '13-Dec',
    'Martin Strupler':      '19-Dec',
    'Vichea Khen':          '22-Dec',
    'Sila Kun':             '27-Dec',
    'Rina Nov':             '01-Jan',
    'Vann Rotha':           '01-Jan',
    'Makara Seng':          '01-Jan',
    'Matthew Seng':         '01-Jan',
    'Savann Sun':           '04-Jan',
    'Florin Flammer':       '07-Jan',
    'Davit Mao':            '07-Jan',
    'Mom Kalic':            '08-Jan',
    'Phat Sout':            '09-Jan',
    'Som Chhang':           '14-Jan',
    'Seav Ey Khean':        '15-Jan',
    'Sotheavy Ngen':        '16-Jan',
    'Sakorun Pok':          '20-Jan',
    'Seyla Nga':            '21-Jan',
    'Rin Nga':              '01-Feb',
    'Sarath Sok':           '01-Feb',
    'Ryna Mom':             '02-Feb',
    'Karano Chhuon':        '03-Feb',
    'Sopheap Heang':        '03-Feb',
    'Nhoem David':          '03-Feb',
    'Sodalin Thangham':     '04-Feb',
    'Nan Houn':             '06-Feb',
    'Sopheaktra Mao':       '07-Feb',
    'Vicheth Song':         '10-Feb',
    'Vutha Kry':            '19-Feb',
    'Saroeun Kim':          '22-Feb',
    'Chrach Chen':          '28-Feb',
    'Kimleang Lun':         '04-Mar',
    'Sothiery Un':          '05-Mar',
    'Kongkea Kouk':         '07-Mar',
    'Parigna Souem':        '07-Mar',
    'Kropumkannika Gnorng': '09-Mar',
    'Tola Khy':             '18-Mar',
    'Aline Barbosa':        '19-Mar',
    'Sreynouch Khit':       '19-Mar',
    'Pech Sang':            '20-Mar',
    'David Piseth':         '28-Mar',
    'Minea Soy':            '28-Mar',
    'Savry Klot':           '01-Apr',
    'Sideth Duch':          '05-Apr',
    'Kanha Un':             '05-Apr',
    'Sreyleak Orn':         '07-Apr',
    'Chanthat Leat':        '08-Apr',
    'Sreymom Thorng':       '08-Apr',
    'Sophy Chek':           '09-Apr',
    'Rachana San':          '10-Apr',
    'Khunhy Thoeun':        '10-Apr',
    'Linet Un':             '10-Apr',
    'Champa Vong':          '14-Apr',
    'Steffi Lendi':         '16-Apr',
    'Kimbuoy Pheng':        '20-Apr',
    'Sopheap Ngat':         '25-Apr',
    'Hing Hoeun':           '01-May',
    'Sreymom Lim':          '02-May',
    'Sotheavy Thea':        '03-May',
    'ND Strupler':          '04-May',
    'Sopheap Puth':         '04-May',
    'Vivian Stumpf':        '06-May',
    'Stephanie Shelow':     '07-May',
    'Noya Men':             '08-May',
    'Sreymom Touch':        '10-May',
    'Vanthen Kim':          '15-May',
    'Sreypeou Hout':        '20-May',
    'Neaksen Sok':          '23-May',
    'Longsamnieng Pol':     '26-May',
    'Ousaphea Lim':         '27-May',
    'Elea Weber':           '28-May',
    'Bethany Roach':        '01-Jun',
    'Hou Sokcheat':         '05-Jun',
    'Bopha Mom':            '05-Jun',
    'Annina Huembeli':      '11-Jun',
    'Nirorn Sok':           '14-Jun',
    'Pichey Ken':           '16-Jun',
    'Matthias Lendi':       '22-Jun',
    'Loy Sambo Hout':       '23-Jun',
    'Rayu Roeun':           '25-Jun',
    'Ya Hoeun':             '28-Jun',
    'Barang Soeurm':        '02-Jul',
    'Sopheak Vat':          '03-Jul',
    'Lieng Huon':           '06-Jul',
    'Sreypich Chhean':      '07-Jul',
    'Srey Roth Outh':       '10-Jul',
    'Kim Sorn Soeuth':      '18-Jul',
}

HIDDEN_STAFF = {"Ryca Mom"}

DEPT_REMAP = {
    'Davit Mao':'Church',
    'Bopha Mom':'Catering','Lieng Huon':'Catering','Som Chhang':'Catering',
    'Phat Sout':'Catering','Pech Sang':'Catering',
    'Vattey Chhun':'Operations','Linet Un':'Operations','Rachana San':'Operations',
    'Kimbuoy Pheng':'Operations','Sakorun Pok':'Operations',
    'Sreymom Lim':'Operations','Kimleang Lun':'Operations',
    'Martin Strupler':'Property',
    'Matthias Lendi':'Social',
    'Stephanie Shelow':'MarCom',
}

DEPT_VALUE_REMAP = {
    'HR':'Human Resources','Fundraising':'Donor Care',
    'Finance':'Operations','IT':'Operations','Admin':'Operations',
    'Family Care':'Social','Medical':'Social','Education':'Social',
    'Daycare':'Social','After School':'Social','Scholarship':'Social',
}

SUB_DEPT_OVERRIDE = {
    # Catering: no sub-dept
    'Bopha Mom':'','Lieng Huon':'','Som Chhang':'','Phat Sout':'','Pech Sang':'',
    # Executive / Founders: no dept or own grouping
    'Sophal Strupler':'',
    # Leadership redistribution: no sub-dept
    'Martin Strupler':'','Matthias Lendi':'','Stephanie Shelow':'',
    # Coffee Shop
    'Sakorun Pok':'Coffee Shop','Sreymom Lim':'Coffee Shop','Kimleang Lun':'Coffee Shop',
    # Property sub-depts
    'Sarath Sok':'Facility Management','Kim Sopheap':'Facility Management',
    'Kim Sorn Soeuth':'Electrical','Chrach Chen':'Electrical',
    'Hing Hoeun':'Security','Hin Loeb':'Security','Ya Hoeun':'Security',
    'Thon Tep':'Security','Saroeun Kim':'Security',
    'Sila Kun':'Cleaning','Hou Sokcheat':'Cleaning','Rin Nga':'Cleaning',
    'Nan Houn':'Gardening','Hean Luos':'Gardening','Minea Soy':'Gardening',
    'Barang Soeurm':'Maintenance','Sopheap Puth':'Maintenance',
    'Sen Meta':'Maintenance','Nhoem David':'Maintenance',
    # Church sub-depts
    'Panha Neang':'Next Steps','Boromey Hom':'Next Steps','Tola Khy':'Next Steps',
    'Bethany Roach':'Outreach','Seangly Leng':'Outreach','Champa Vong':'Outreach',
    'Rothana Ros':'Outreach','Sambath Song':'Outreach','Noa Flammer':'Outreach',
    'Ranaomi Heak':'Outreach',
    'Sombo Ros':'Worship','Sreileak Rous':'Worship','Dana Thy':'Worship','Pichey Ken':'Worship',
    'Mom Kalic':'Worship & Technical','Kanal Tensok':'Worship & Technical',
    'Somun Yem':'Worship & Technical','Elea Weber':'Worship & Technical',
    'Sreypich Chhean':'Youth','Sotheavy Ngen':'Youth',
    'Ryna Mom':'Kids','Khemry Men':'Kids','Robson Barbosa':'Kids','Aline Barbosa':'Kids',
    'Sovansreyreach Tat':'Kids','Sopheap Ngat':'Kids','Sopheak Vat':'Kids',
    # Social
    'Chintha Pav':'After School',
    # Donor Care
    'Rebecca Vey':'Hospitality',
}

ORG = {
    "ND Strupler":{"manager":None},"Sophal Strupler":{"manager":"ND Strupler"},
    "Longsamnieng Pol":{"manager":"ND Strupler"},"Steffi Lendi":{"manager":"ND Strupler"},
    "Annina Huembeli":{"manager":"ND Strupler"},
    "Eddie Roach":{"manager":None},
    "Rany Mom":{"manager":"Eddie Roach"},"Davit Mao":{"manager":"Rany Mom"},
    "Mom Kalic":{"manager":"Eddie Roach"},"Kanal Tensok":{"manager":"Mom Kalic"},
    "Somun Yem":{"manager":"Mom Kalic"},"Elea Weber":{"manager":"Mom Kalic"},
    "Sombo Ros":{"manager":"Eddie Roach"},"Sreileak Rous":{"manager":"Sombo Ros"},
    "Dana Thy":{"manager":"Sombo Ros"},"Pichey Ken":{"manager":"Sombo Ros"},
    "Bethany Roach":{"manager":"Eddie Roach"},"Ranaomi Heak":{"manager":"Bethany Roach"},
    "Noa Flammer":{"manager":"Bethany Roach"},"Seangly Leng":{"manager":"Bethany Roach"},
    "Champa Vong":{"manager":"Bethany Roach"},"Rothana Ros":{"manager":"Bethany Roach"},
    "Sambath Song":{"manager":"Bethany Roach"},
    "Panha Neang":{"manager":"Eddie Roach"},"Boromey Hom":{"manager":"Panha Neang"},
    "Sreypich Chhean":{"manager":"Panha Neang"},"Sotheavy Ngen":{"manager":"Sreypich Chhean"},
    "Tola Khy":{"manager":"Panha Neang"},
    "Ryna Mom":{"manager":"Eddie Roach"},"Khemry Men":{"manager":"Ryna Mom"},
    "Sopheap Ngat":{"manager":"Ryna Mom"},"Sopheak Vat":{"manager":"Ryna Mom"},
    "Sovansreyreach Tat":{"manager":"Ryna Mom"},
    "Robson Barbosa":{"manager":"Eddie Roach"},"Aline Barbosa":{"manager":"Robson Barbosa"},
    "Simone Strupler":{"manager":None},"Sothiery Un":{"manager":"Simone Strupler"},
    "Vann Rotha":{"manager":"Simone Strupler"},"Kropumkannika Gnorng":{"manager":"Simone Strupler"},
    "Matthias Lendi":{"manager":None},
    "Parigna Souem":{"manager":"Matthias Lendi"},"Khunhy Thoeun":{"manager":"Matthias Lendi"},
    "Su Sey":{"manager":"Khunhy Thoeun"},"Nirorn Sok":{"manager":"Khunhy Thoeun"},
    "David Piseth":{"manager":"Khunhy Thoeun"},"Sreynouch Khit":{"manager":"Khunhy Thoeun"},
    "Sotheany Ngeun":{"manager":"Khunhy Thoeun"},"Tola Khem":{"manager":"Khunhy Thoeun"},
    "Vicheth Song":{"manager":"Khunhy Thoeun"},"Seav Ey Khean":{"manager":"Khunhy Thoeun"},
    "Phin Teb":{"manager":"Matthias Lendi"},"Makara Seng":{"manager":"Phin Teb"},
    "Srytom Pogn":{"manager":"Phin Teb"},
    "Thavy Tham":{"manager":"Matthias Lendi"},"Nuthida Han":{"manager":"Thavy Tham"},
    "Sodalin Thangham":{"manager":"Thavy Tham"},"Vivian Stumpf":{"manager":"Thavy Tham"},
    "Bopha Mom":{"manager":"Thavy Tham"},"Lieng Huon":{"manager":"Bopha Mom"},
    "Som Chhang":{"manager":"Bopha Mom"},"Phat Sout":{"manager":"Bopha Mom"},
    "Pech Sang":{"manager":"Bopha Mom"},
    "Karano Chhuon":{"manager":None},
    "Sreymom Thorng":{"manager":"Karano Chhuon"},"Sopheap Heang":{"manager":"Karano Chhuon"},
    "Sreymom Touch":{"manager":"Karano Chhuon"},"Seyla Nga":{"manager":"Karano Chhuon"},
    "Kolap Touch":{"manager":"Karano Chhuon"},"Vichea Khen":{"manager":"Karano Chhuon"},
    "Chanthat Leat":{"manager":"Karano Chhuon"},"Sotheavy Thea":{"manager":"Karano Chhuon"},
    "Rongroeung Chamroeun":{"manager":"Karano Chhuon"},"Champa Toch":{"manager":"Karano Chhuon"},
    "Sreypeou Hout":{"manager":"Karano Chhuon"},"Srey Roth Outh":{"manager":"Karano Chhuon"},
    "Sopheaktra Mao":{"manager":"Karano Chhuon"},"Sokuntheary Sim":{"manager":"Karano Chhuon"},
    "Sideth Duch":{"manager":"Karano Chhuon"},"Chamroeun Muon":{"manager":"Karano Chhuon"},
    "Savry Klot":{"manager":"Karano Chhuon"},"Sophea Mak":{"manager":"Karano Chhuon"},
    "Noya Men":{"manager":"Karano Chhuon"},"Thea Rith Keo":{"manager":"Karano Chhuon"},
    "Sa Soueyvan":{"manager":"Karano Chhuon"},"Matthew Seng":{"manager":"Karano Chhuon"},
    "Sophy Chek":{"manager":"Karano Chhuon"},"Hana Seap":{"manager":"Karano Chhuon"},
    "Vattey Chhun":{"manager":None},
    "Kanha Un":{"manager":"Vattey Chhun"},"Linet Un":{"manager":"Vattey Chhun"},
    "Rachana San":{"manager":"Linet Un"},"Kimbuoy Pheng":{"manager":"Linet Un"},
    "Vanthen Kim":{"manager":"Vattey Chhun"},"Makara Tes":{"manager":"Vanthen Kim"},
    "Neaksen Sok":{"manager":"Vanthen Kim"},
    "Sakorun Pok":{"manager":"Vattey Chhun"},"Sreymom Lim":{"manager":"Sakorun Pok"},
    "Kimleang Lun":{"manager":"Sakorun Pok"},
    "Stephanie Shelow":{"manager":None},
    "Ratana Khy":{"manager":"Stephanie Shelow"},"Florin Flammer":{"manager":"Ratana Khy"},
    "Sengly Muol":{"manager":"Ratana Khy"},"Soktheavy Tann":{"manager":"Ratana Khy"},
    "Vutha Kry":{"manager":"Ratana Khy"},"Savann Sun":{"manager":"Ratana Khy"},
    "Rayu Roeun":{"manager":"Ratana Khy"},
    "Martin Strupler":{"manager":None},
    "Sarath Sok":{"manager":"Martin Strupler"},"Kim Sopheap":{"manager":"Sarath Sok"},
    "Hing Hoeun":{"manager":"Sarath Sok"},"Hin Loeb":{"manager":"Hing Hoeun"},
    "Ya Hoeun":{"manager":"Hing Hoeun"},"Thon Tep":{"manager":"Hing Hoeun"},
    "Saroeun Kim":{"manager":"Hing Hoeun"},
    "Sila Kun":{"manager":"Sarath Sok"},"Hou Sokcheat":{"manager":"Sila Kun"},
    "Rin Nga":{"manager":"Sila Kun"},
    "Nan Houn":{"manager":"Sarath Sok"},"Hean Luos":{"manager":"Nan Houn"},
    "Minea Soy":{"manager":"Nan Houn"},
    "Kim Sorn Soeuth":{"manager":"Sarath Sok"},"Chrach Chen":{"manager":"Kim Sorn Soeuth"},
    "Barang Soeurm":{"manager":"Sarath Sok"},"Sopheap Puth":{"manager":"Barang Soeurm"},
    "Sen Meta":{"manager":"Barang Soeurm"},"Nhoem David":{"manager":"Barang Soeurm"},
    "Kongkea Kouk":{"manager":"Longsamnieng Pol"},"Jane Meas":{"manager":"Longsamnieng Pol"},
    "Sreyleak Orn":{"manager":"Longsamnieng Pol"},
    "Loy Sambo Hout":{"manager":"Steffi Lendi"},"Seavva Han":{"manager":"Steffi Lendi"},
}

ROLE_OVERRIDE = {'ND Strupler': 'Founder'}

ALIASES = {'Longsamnieng Pol':'Paul','Seavva Han':'Sophie'}

STATIC_STAFF = [
    {'name':'Sopheak Phol','role':'University Scholarship Student','dept_raw':'Scholarship','email':'sopheak.phol@icf-cambodia.com','phone':'+855967093178','telegram':'Sopheakphol'},
    {'name':'Ry Rey','role':'University Scholarship Student','dept_raw':'Scholarship','email':'ry.rey@icf-cambodia.com','phone':'+855889314500','telegram':''},
    {'name':'Sinat Hoem','role':'University Scholarship Student','dept_raw':'Scholarship','email':'sinat.hoem@icf-cambodia.com','phone':'','telegram':'Sintna12345'},
    {'name':'Sreynuch Phuk','role':'University Scholarship Student','dept_raw':'Scholarship','email':'sreynuch.phuk@icf-cambodia.com','phone':'+85587998079','telegram':'SreynuchPhuk'},
    {'name':'Prerk Sey','role':'University Scholarship Student','dept_raw':'Scholarship','email':'prerk.sey@icf-cambodia.com','phone':'','telegram':''},
    {'name':'Kanha Vat','role':'University Scholarship Student','dept_raw':'Scholarship','email':'kanha.vat@icf-cambodia.com','phone':'+855888603145','telegram':''},
    {'name':'Sreylin Nan','role':'University Scholarship Student','dept_raw':'Scholarship','email':'sreylin.nan@icf-cambodia.com','phone':'+85517501821','telegram':''},
    {'name':'Makara Chy','role':'University Scholarship Student','dept_raw':'Scholarship','email':'makara.chy@icf-cambodia.com','phone':'+855972705807','telegram':'CHY_MAKARA290106'},
    {'name':'Sriydav Nak','role':'University Scholarship Student','dept_raw':'Scholarship','email':'sriydav.nak@icf-cambodia.com','phone':'+855716489271','telegram':'SreydavNak'},
    {'name':'Monyroth Ev','role':'University Scholarship Student','dept_raw':'Scholarship','email':'monyroth.ev@icf-cambodia.com','phone':'+85560293074','telegram':'Nyrorth'},
    {'name':'Sreyleak Houn','role':'University Scholarship Student','dept_raw':'Scholarship','email':'sreyleak.houn@icf-cambodia.com','phone':'+855719608193','telegram':'Sreyleakhounn'},
    {'name':'Sary Sarin','role':'University Scholarship Student','dept_raw':'Scholarship','email':'sary.sarin@icf-cambodia.com','phone':'','telegram':'Sary_sarin'},
    {'name':'Sothea Thov','role':'University Scholarship Student','dept_raw':'Scholarship','email':'sothea.thov@icf-cambodia.com','phone':'','telegram':'SotheaThov'},
    {'name':'Chintha Pav','role':'University Scholarship Student','dept_raw':'Social','email':'chintha.pav@icf-cambodia.com','phone':'+855975218278','telegram':'barvchenda17'},
    {'name':'Chhaivorn Vin','role':'University Scholarship Student','dept_raw':'Scholarship','email':'chhaivorn.vin@icf-cambodia.com','phone':'','telegram':'chhaivor'},
    {'name':'Naroth Roth','role':'University Scholarship Student','dept_raw':'Scholarship','email':'naroth.roth@icf-cambodia.com','phone':'','telegram':'RothNaroth'},
    {'name':'Hun Hub','role':'University Scholarship Student','dept_raw':'Scholarship','email':'hun.hub@icf-cambodia.com','phone':'','telegram':''},
    {'name':'Rebecca Vey','role':'University Scholarship Student','dept_raw':'Donor Care','email':'rebecca.vey@icf-cambodia.com','phone':'+85595825470','telegram':'rebeccavey'},
    {'name':'Sovann Choeut','role':'University Scholarship Student','dept_raw':'Scholarship','email':'sovann.choeut@icf-cambodia.com','phone':'+85511323611','telegram':'choeutsovann'},
]

# ── Parse CSV ─────────────────────────────────────────────────────────────────
with open('/sessions/vibrant-keen-volta/mnt/outputs/staff_contacts.csv','r',encoding='utf-8') as f:
    rows = list(csv.reader(f))

header = rows[0]
col = {}
for i,h in enumerate(header):
    hl = h.strip().lower()
    if '#' in hl: col['num']=i
    elif 'name' in hl: col['name']=i
    elif 'role' in hl: col['role']=i
    elif 'department' in hl or hl=='dept': col['dept']=i
    elif 'email' in hl: col['email']=i
    elif 'phone' in hl: col['phone']=i
    elif 'telegram' in hl: col['telegram']=i
    elif 'photo' in hl: col['photo']=i

contacts = []
for r in rows[1:]:
    if not r or all(not c.strip() for c in r): continue
    name = r[col['name']].strip() if 'name' in col else ''
    if not name or name in HIDDEN_STAFF: continue
    dept_raw = r[col['dept']].strip() if 'dept' in col else ''
    dept_display = DEPT_REMAP.get(name) or DEPT_VALUE_REMAP.get(dept_raw) or dept_raw
    if name in SUB_DEPT_OVERRIDE:
        sub_dept = SUB_DEPT_OVERRIDE[name]
    elif name in DEPT_REMAP or dept_raw in DEPT_VALUE_REMAP:
        sub_dept = dept_raw
    else:
        sub_dept = ''
    photo_raw = r[col['photo']].strip() if 'photo' in col else ''
    photo_url = photo_raw
    if 'drive.google.com' in photo_raw:
        m = re.search(r'[?&]id=([^&]+)',photo_raw) or re.search(r'/file/d/([^/]+)',photo_raw)
        if m: photo_url = f'https://lh3.googleusercontent.com/d/{m.group(1)}=s400'
    role_raw = r[col['role']].strip() if 'role' in col else ''
    contacts.append({
        'name':name,'role':ROLE_OVERRIDE.get(name, role_raw),
        'dept':dept_display,'sub_dept':sub_dept,
        'reporting_to':ORG.get(name,{}).get('manager') or '',
        'email':r[col['email']].strip() if 'email' in col else '',
        'phone':r[col['phone']].strip() if 'phone' in col else '',
        'telegram':r[col['telegram']].strip() if 'telegram' in col else '',
        'photo_url':photo_url,'alias':ALIASES.get(name,''),
        'birthday':BIRTHDAYS.get(name,''),
        'start_date':START_DATES.get(name,''),
        'source':'sheet',
    })

sheet_names = {c['name'] for c in contacts}
for s in STATIC_STAFF:
    if s['name'] in sheet_names: continue
    raw = s['dept_raw']
    dept_display = raw  # Use raw directly — Scholarship stays Scholarship, not remapped to Social
    contacts.append({
        'name':s['name'],'role':s['role'],'dept':dept_display,
        'sub_dept':SUB_DEPT_OVERRIDE.get(s['name'],''),
        'reporting_to':ORG.get(s['name'],{}).get('manager') or '',
        'email':s['email'],'phone':s['phone'],'telegram':s['telegram'],
        'photo_url':'','alias':'',
        'birthday':BIRTHDAYS.get(s['name'],''),
        'start_date':START_DATES.get(s['name'],''),
        'source':'static',
    })

contacts.sort(key=lambda x: x['name'].lower())
print(f"Total: {len(contacts)}  |  Birthdays: {sum(1 for c in contacts if c['birthday'])}  |  Start dates: {sum(1 for c in contacts if c['start_date'])}")
dept_counts = Counter(c['dept'] for c in contacts)
for d,n in sorted(dept_counts.items()): print(f"  {d}: {n}")

# ── Style helpers ──────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill('solid',fgColor='1F2937')
HEADER_FONT = Font(name='Arial',bold=True,color='FFFFFF',size=10)
GROUP_FILL  = PatternFill('solid',fgColor='CFE2F3')
GROUP_FONT  = Font(name='Arial',bold=True,color='1A3A5C',size=10)
DATA_FONT   = Font(name='Arial',size=10)
GRAY_FONT   = Font(name='Arial',size=10,italic=True,color='6B7280')
THIN = Border(
    left=Side(style='thin',color='D0D0D0'),right=Side(style='thin',color='D0D0D0'),
    top=Side(style='thin',color='D0D0D0'),bottom=Side(style='thin',color='D0D0D0'),
)
CENTER = Alignment(horizontal='center',vertical='center')
LEFT   = Alignment(horizontal='left',vertical='center')

COLUMNS = [
    ('#',4),('Name',24),('Role',28),('Department',16),('Sub-Dept',16),
    ('Reporting To',22),('Email',32),('Phone',16),('Telegram',16),
    ('Birthday',10),('Start Date',12),('End Date',12),('Alias',13),('Photo URL',40),
]

def write_header(ws):
    for ci,(lbl,wid) in enumerate(COLUMNS,1):
        cell = ws.cell(row=1,column=ci,value=lbl)
        cell.fill=HEADER_FILL; cell.font=HEADER_FONT
        cell.alignment=CENTER; cell.border=THIN
        ws.column_dimensions[get_column_letter(ci)].width=wid
    ws.row_dimensions[1].height=22
    ws.freeze_panes='B2'

def write_group_row(ws, row_num, label, n_staff):
    ws.row_dimensions[row_num].height=18
    ws.merge_cells(start_row=row_num,start_column=1,end_row=row_num,end_column=len(COLUMNS))
    cell = ws.cell(row=row_num,column=1,value=f'  {label}  ·  {n_staff} staff')
    cell.fill=GROUP_FILL; cell.font=GROUP_FONT
    cell.alignment=LEFT; cell.border=THIN

def write_person_row(ws, row_num, num, p):
    ws.row_dimensions[row_num].height=16
    font = GRAY_FONT if p['source']=='static' else DATA_FONT
    vals = [num,p['name'],p['role'],p['dept'],p['sub_dept'],p['reporting_to'],
            p['email'],p['phone'],p['telegram'],p['birthday'],p['start_date'],'',p['alias'],p['photo_url']]
    for ci,val in enumerate(vals,1):
        cell = ws.cell(row=row_num,column=ci,value=val)
        cell.font=font
        cell.alignment=CENTER if ci==1 else LEFT
        cell.border=THIN

# ── Groupings ──────────────────────────────────────────────────────────────────
DEPT_ORDER = ['Church','Catering','Donor Care','Guest Relations','Human Resources',
              'MarCom','Operations','Property','Social','Scholarship']

dept_groups = defaultdict(list)
for c in contacts:
    dept_groups[c['dept'] or 'Other'].append(c)

# Departments that get their own sheet (match dept pages)
DEPT_SHEETS = ['Church','Catering','Donor Care','Guest Relations','Human Resources',
               'MarCom','Operations','Property','Social','Scholarship']

# ── Build workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Sheet 1: All Staff ─────────────────────────────────────────────────────────
ws_all = wb.active
ws_all.title = 'All Staff'
write_header(ws_all)

row_num = 2
grand_num = 1
ordered_depts = [d for d in DEPT_ORDER if d in dept_groups]
ordered_depts += sorted(d for d in dept_groups if d not in DEPT_ORDER)

for dept in ordered_depts:
    group = dept_groups[dept]
    write_group_row(ws_all, row_num, dept, len(group))
    row_num += 1
    for p in group:
        write_person_row(ws_all, row_num, grand_num, p)
        grand_num += 1
        row_num += 1

# ── Department sheets ──────────────────────────────────────────────────────────
# Dept-specific sheet columns (no Department col, use Sub-Dept as grouping)
DEPT_COLS = [
    ('#',4),('Name',24),('Role',30),('Sub-Dept',18),
    ('Reporting To',22),('Email',32),('Phone',16),('Telegram',16),
    ('Birthday',10),('Start Date',12),('End Date',12),('Alias',13),('Photo URL',40),
]

# Dept sheet header colors per dept
DEPT_COLORS = {
    'Church':         ('5B21B6','FFFFFF'),
    'Catering':       ('92400E','FFFFFF'),
    'Donor Care':     ('065F46','FFFFFF'),
    'Guest Relations':('0E4C6E','FFFFFF'),
    'Human Resources':('1E3A5F','FFFFFF'),
    'MarCom':         ('7C2D12','FFFFFF'),
    'Operations':     ('1E3A5F','FFFFFF'),
    'Property':       ('92400E','FFFFFF'),
    'Social':         ('1A4731','FFFFFF'),
    'Scholarship':    ('374151','FFFFFF'),
}

def write_dept_header(ws, dept):
    bg, fg = DEPT_COLORS.get(dept, ('1F2937','FFFFFF'))
    hfill = PatternFill('solid',fgColor=bg)
    hfont = Font(name='Arial',bold=True,color=fg,size=10)
    for ci,(lbl,wid) in enumerate(DEPT_COLS,1):
        cell = ws.cell(row=1,column=ci,value=lbl)
        cell.fill=hfill; cell.font=hfont
        cell.alignment=CENTER; cell.border=THIN
        ws.column_dimensions[get_column_letter(ci)].width=wid
    ws.row_dimensions[1].height=22
    ws.freeze_panes='B2'

def write_dept_group_row(ws, row_num, label, n_staff, dept):
    bg, _ = DEPT_COLORS.get(dept, ('CFE2F3','1A3A5C'))
    # Lighten the dept color for group rows
    light_fill = PatternFill('solid',fgColor='CFE2F3')
    ws.row_dimensions[row_num].height=18
    ws.merge_cells(start_row=row_num,start_column=1,end_row=row_num,end_column=len(DEPT_COLS))
    cell = ws.cell(row=row_num,column=1,value=f'  {label}  ·  {n_staff} staff')
    cell.fill=light_fill; cell.font=GROUP_FONT
    cell.alignment=LEFT; cell.border=THIN

def write_dept_person_row(ws, row_num, num, p):
    ws.row_dimensions[row_num].height=16
    font = GRAY_FONT if p['source']=='static' else DATA_FONT
    vals = [num,p['name'],p['role'],p['sub_dept'],p['reporting_to'],
            p['email'],p['phone'],p['telegram'],p['birthday'],p['start_date'],'',p['alias'],p['photo_url']]
    for ci,val in enumerate(vals,1):
        cell = ws.cell(row=row_num,column=ci,value=val)
        cell.font=font
        cell.alignment=CENTER if ci==1 else LEFT
        cell.border=THIN

for dept in DEPT_SHEETS:
    if dept not in dept_groups: continue
    group = dept_groups[dept]
    ws = wb.create_sheet(dept)
    write_dept_header(ws, dept)

    # Group by sub-dept within this dept sheet
    sub_groups = defaultdict(list)
    for p in group:
        sub_groups[p['sub_dept'] or '—'].append(p)

    row_num = 2
    local_num = 1
    # Order sub-depts: named ones first, then '—'
    named_subs = [s for s in sub_groups if s != '—']
    named_subs.sort()
    sub_order = named_subs + (['—'] if '—' in sub_groups else [])

    for sub in sub_order:
        sub_people = sub_groups[sub]
        if len(named_subs) > 0:  # Only show sub-group headers if there are sub-depts
            write_dept_group_row(ws, row_num, sub if sub != '—' else dept, len(sub_people), dept)
            row_num += 1
        for p in sub_people:
            write_dept_person_row(ws, row_num, local_num, p)
            local_num += 1
            row_num += 1

# ── Position History sheet ─────────────────────────────────────────────────────
ws_hist = wb.create_sheet('Position History')
hist_cols = [
    ('Date',12),('Name',24),('Previous Role',28),('New Role',28),
    ('Department',18),('Sub-Dept',20),('Reporting To',22),('Notes',40),
]
for ci,(lbl,wid) in enumerate(hist_cols,1):
    cell = ws_hist.cell(row=1,column=ci,value=lbl)
    cell.fill=HEADER_FILL; cell.font=HEADER_FONT
    cell.alignment=CENTER; cell.border=THIN
    ws_hist.column_dimensions[get_column_letter(ci)].width=wid
ws_hist.row_dimensions[1].height=22
ws_hist.freeze_panes='A2'

# Boromey Hom — scheduled position change (Aug 20, 2026)
YELLOW_FILL = PatternFill('solid',fgColor='FFF9C4')
ORANGE_FILL = PatternFill('solid',fgColor='FEF3C7')
boromey = ['2026-08-20','Boromey Hom','Next Steps Facilitator','Program Specialist',
           'Church','Worship & Technical','Mom Kalic','Scheduled — position change effective 20 Aug 2026']
for ci,v in enumerate(boromey,1):
    cell = ws_hist.cell(row=2,column=ci,value=v)
    cell.font=Font(name='Arial',size=10,bold=(ci in (2,3,4)))
    cell.fill=ORANGE_FILL; cell.alignment=LEFT; cell.border=THIN
ws_hist.row_dimensions[2].height=16

out = '/sessions/vibrant-keen-volta/mnt/ICF-Staff-Hub/Staff_Contacts_Final.xlsx'
wb.save(out)
print(f"\nSaved → {out}")
